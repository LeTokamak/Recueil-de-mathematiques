import openpyxl as opxl
import csv

# %% Couleurs

noir         = "00000000" #000000
rouge        = "00FF0000" #FF0000
bleu         = "000000CC" #0000CC
vert_clair   = "00CCFF99" #CCFF99
rouge_clair  = "00FF9999" #FF9999
jaune_clair  = "00FFFF99" #FFFF99
gris_clair   = "00C0C0C0" #C0C0C0
blanc        = "00FFFFFF" #FFFFFF

fond_entete  = "00222222" #222222
fond_degre   = "00666666" #666666
fond_rad     = "00DDDDDD" #AAAAAA #"00AAAAAA" 
fond_colonne = "00DDDDDD" #DDDDDD
fond_blanc   = "00FFFFFF" #FFFFFF



def ouvrir_fichier_excel(lien_fichier_excel) :
    """
    Cette fonction ne supporte que les formats suivant : 
    .xlsx, .xlsm, .xltx, .xltm
    """
    wb = opxl.load_workbook(filename = lien_fichier_excel)
    return wb

def ouvrir_fichier_csv(lien_fichier_csv, separateur = ";") :
    """
    Cette fonction ne supporte que le format suivant : 
    .csv
    """
    liste_lignes = []
    
    with open(lien_fichier_csv, newline='') as csvfile:
        for row in csv.reader(csvfile, delimiter = separateur, quotechar='|'):
            liste_lignes.append(row)
            
    return liste_lignes



def creation_fichier_excel():
    wb = opxl.Workbook()
    return wb



def police( taille          = 11, 
            est_en_gras     = False, 
            est_en_italique = False, 
            est_souligne    = False, 
            couleur         = noir,
            police          = "Calibri"):
    
    if est_souligne : est_souligne = ("singleAccounting", "doubleAccounting", "double", "single")[-1]
    else            : est_souligne =  "none"
    
    font = opxl.styles.Font( name      = police         ,
                             size      = taille         ,
                             bold      = est_en_gras    ,
                             italic    = est_en_italique,
                             underline = est_souligne   ,
                             color     = couleur         )
    
    return font



def couleur_fond(couleur = blanc):
    
    fond = opxl.styles.PatternFill(start_color = couleur,
                                   end_color   = couleur,
                                   fill_type   = 'solid')
    
    return fond



def alignement(alignement_horizontal="c", alignement_vertical="c", angle_texte = 0):
    
    liste_horizontale = ("fill", "general", "justify", "center", "left", "right"  , "centerContinuous", "distributed" )
    liste_verticale   = (                   "justify", "center", "top" , "bottom" ,                     "distributed" )
    
    if alignement_horizontal in ("g", "l", "gauche", "left"  ) : position_horizontale = liste_horizontale[4]
    if alignement_vertical   in ("h", "t", "haut"  , "top"   ) : position_verticale   = liste_verticale  [2]
    
    if alignement_horizontal in ("c",      "centre", "center") : position_horizontale = liste_horizontale[3]
    if alignement_vertical   in ("c",      "centre", "center") : position_verticale   = liste_verticale  [1]
    
    if alignement_horizontal in ("d", "r", "droite", "right" ) : position_horizontale = liste_horizontale[5]
    if alignement_vertical   in ("b",      "bas"   , "bottom") : position_verticale   = liste_verticale  [3]
    
    return opxl.styles.Alignment(horizontal   = position_horizontale, 
                                 vertical     = position_verticale  , 
                                 textRotation = angle_texte          )


def commentaire ( texte           = "",
                  couleur_de_fond = jaune_clair, 
                  auteur          = "par_defaut",
                  largeur         = 700,
                  hauteur         = 200  ):
    
    commentaire = opxl.comments.Comment(texte, author = auteur)
    commentaire.width = largeur     # Largeur du commentaire en pixels
    commentaire.height = hauteur    # Hauteur du commentaire en pixels
    # commentaire.width = 200     # Largeur du commentaire en pixels
    # commentaire.height = 100    # Hauteur du commentaire en pixels
    # commentaire.visible = True  # Afficher le commentaire
    # commentaire.anchor = 'E'    # Ancrer le commentaire à droite de la cellule
    # commentaire.font = openpyxl.styles.Font(name='Arial', size=12, italic=True)
    commentaire.fill = couleur_fond(couleur_de_fond)
    
    return commentaire

def redimensionnement_colonne_auto_optimal(feuille):
    """
    LibreOffice : 1,57 cm = 8"""
    dims = {}
    for row in feuille.rows :
        for cell in row :
            if cell.value : 
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))    
    
    for col, value in dims.items() :
        print(opxl.utils.get_column_letter(col), end=" - ")
        print(value+1)
        feuille.column_dimensions[opxl.utils.get_column_letter(col)].width = value+1
        #feuille.column_dimensions[opxl.utils.get_column_letter(col)].auto_size = True
    """
    for colonne in feuille.columns :
        colonne.
    """
    
def bordure( style = "thin", couleur = noir ,a_droite = False, a_gauche = False, en_haut = False, en_bas = False):
    """
    style = {
        "dashDot", 
        "mediumDashDotDot", 
        "double", 
        "mediumDashed", 
        "dashDotDot", 
        "medium",
        "thick",
        "dashed", 
        "mediumDashDot", 
        "dotted", 
        "hair",
        "slantDashDot",
        "thin"
    }
    """
    bordure = opxl.styles.Side(style=style, color=couleur)
    
    if a_droite : return opxl.styles.Border(right=bordure)
    if a_gauche : return opxl.styles.Border(left=bordure)
    if en_haut  : return opxl.styles.Border(top=bordure)
    if en_bas   : return opxl.styles.Border(bottom=bordure)
    
    if a_droite and a_gauche : return opxl.styles.Border(right=bordure, left=bordure)
    if a_droite and en_haut  : return opxl.styles.Border(right=bordure, top=bordure)
    if a_droite and en_bas   : return opxl.styles.Border(right=bordure, bottom=bordure)
    if a_gauche and en_haut  : return opxl.styles.Border(left=bordure, top=bordure)
    if a_gauche and en_bas   : return opxl.styles.Border(left=bordure, bottom=bordure)
    if en_haut  and en_bas   : return opxl.styles.Border(top=bordure, bottom=bordure)
    
    if a_droite and a_gauche and en_haut : return opxl.styles.Border(right=bordure, left=bordure, top=bordure)
    if a_droite and a_gauche and en_bas  : return opxl.styles.Border(right=bordure, left=bordure, bottom=bordure)
    if a_droite and en_haut  and en_bas  : return opxl.styles.Border(right=bordure, top=bordure, bottom=bordure)
    if a_gauche and en_haut  and en_bas  : return opxl.styles.Border(left=bordure, top=bordure, bottom=bordure)
    
    if a_droite and a_gauche and en_haut and en_bas : return opxl.styles.Border(right=bordure, left=bordure, top=bordure, bottom=bordure)
    


"""


# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
"""