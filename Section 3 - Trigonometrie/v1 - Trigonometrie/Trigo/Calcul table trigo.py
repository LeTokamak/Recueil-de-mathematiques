from openpyxl import Workbook
import math as m

wb = Workbook()
ws = wb.active

derNbPremier = 2000

# %% Mise en Page

Intro = ["Degré", "Rad", None, None, "sin (x)", "cos (x)", "tan (x)", None,  "log |sin (x)|", "log |cos (x)|", "log |tan (x)|"]

ws.append(Intro)
ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)

# %% Fonctions

def tronque(n, nbDecimal):
    """
    Coupe un réel n avec nbDecimal chiffres après la virgule
    """
    return float(int(n*(10**nbDecimal)))/(10**nbDecimal)

def tronqueEntier(x, nbChiffre):
    """
    Coupe un réel x avec nbChiffre chiffres significatifs
    """    
    return int(x/(10**nbChiffre))*(10**nbChiffre)
   
def FonctArrondie(n, fonct, fonctReciproque, nbDecimal):
    """
    Renvoie la valeur de la fonction fonct(n) arrondie à nbDecimal près
    Cet arrondi est le plus proche possible de la valeur exacte
    """
    try :
        yExact = fonct(n)
        yMin = tronque(yExact, nbDecimal)
        yMaj = yMin + 10**(-1 * nbDecimal)
    
        if fonctReciproque(yMin) - n < fonctReciproque(yMaj) - n :
            return yMin
        else :
            return yMaj
    except : 
        return 404


def MiseEnPage(nb, nbChiffre, nbDecimal, Espace = True, affichageSigne = False, retourneEntier = False):
    """
    Met en page un nombre en : 
        - ajoutant des 0 devant le nombre pour que le nb de chiffres devant la virgule soit égal à nbChiffre
        - ajoutant des 0 après la virgule pour que le nb de chiffres après la virgule soit égal à nbDecimal
        - ajoutant des espaces de séparation entre les groupes de chiffres (Espace = True)
        - ajoutant un signe devant le nombre (affichageSigne = True)
        - retournant un nombre entier (retourneEntier = True)
    """
    if "e" in str(nb) : 
        return str(nb)
    
    x  = round(nb, nbDecimal)
    nb = abs(nb)
    
    if Espace :
        # Détermination du nombre de 0 à écrire devant le nombre
        nbZeroAv = nbChiffre - len(str(int(nb)))
        if nbZeroAv < 0 :
            return "ERREUR ! nbZeroAv < 0"
        
        # Détermination de la taille du groupe de poids le plus fort (en se basant sur la convention d'écriture "12 054")
        nbChiffreGroupeAv = nbChiffre % 3
        if nbChiffreGroupeAv == 0 :
            nbChiffreGroupeAv = 3 

        # Écriture du nombre avec ses espaces de séparation
        strNbInt = nbZeroAv * "0" + str(int(nb))
        strNbFinal = strNbInt[:nbChiffreGroupeAv] 
        if len(str(int(nb))) != nbChiffreGroupeAv :
            strNbFinal += " "
        
        for i in range(len(strNbInt[nbChiffreGroupeAv:]) ):
            strNbFinal += strNbInt[nbChiffreGroupeAv + i]
            if (i+1) % 3 == 0 and i != len(strNbInt[nbChiffreGroupeAv:])-1:
                strNbFinal += " "
            
        if not retourneEntier :
            strNbFinal += "."
           
            decimales = str(nb)[len(str(int(nb)))+1:]
            
            for i in range (nbDecimal):
                if i < len(decimales) :
                    if i % 3 == 0 and i != 0:
                        strNbFinal += " "
                    strNbFinal += decimales[i]
                else :
                    strNbFinal += "0"
    
    else :
        strNbFinal = nbZeroAv * "0" + str(nb) + nbDecimal * "0"
        
    if affichageSigne and x < 0:
        strNbFinal = "- " + strNbFinal
    
    return strNbFinal
        
def EstPremier(n):
    """
    Renvoie True si n est premier, False sinon
    """
    d = 2
    while d*d <= n :
        if n % d == 0 :
            return False
        d += 1
    return True

def ListeNbPremierJusqua(nMax):
    """
    Renvoie la liste des nombres premiers jusqu'à nMax
    """
    n = 2
    L = []
    
    while n < nMax :
        if EstPremier(n):
            L.append(n)
        n += 1
    return L

LPremier = ListeNbPremierJusqua(derNbPremier)

def Simplification(n,d):
    """
    Simplifie une fraction n/d
    Renvoie une chaîne de caractère, le numérateur et le dénominateur simplifiés, les facteurs premiers du numérateur et du dénominateur
    """
    if n == 0 : return "0",1,1,1,1 ;    
    
    nFact = []
    nNbTest = LPremier[:].copy()
    dFact = []
    dNbTest = LPremier[:].copy()
    
    while len(nNbTest) != 0 : # Nominateur
        for i in nNbTest:
            if n % i == 0 :
                nFact.append(i)
                n = n / i
            else :
                nNbTest.remove(i)
                
    while len(dNbTest) != 0 : # Dénominateur
        for i in dNbTest:
            if d % i == 0 :
                dFact.append(i)
                d = d / i
            else :
                dNbTest.remove(i)
                
    for i in nFact :
        if i in dFact:
            nFact.remove(i)
            dFact.remove(i)

    for i in dFact :
        if i in nFact:
            nFact.remove(i)
            dFact.remove(i)
    
    for i in nFact :
        if i in dFact:
            nFact.remove(i)
            dFact.remove(i)

    for i in dFact :
        if i in nFact:
            nFact.remove(i)
            dFact.remove(i)
    
    nomi,deno = 1,1
    
    for i in nFact:
        nomi = nomi * i

    for i in dFact:
        deno = deno * i
    
    if nomi == 1 :
        strNomi = ""
    else :
        strNomi = str(nomi)
        
    if n < 0 :
        signe = "- "
    else :
        signe = ""
    
    if nomi == deno :
        strFract = signe + "π"
    else : 
        strFract = signe + strNomi + " π / " + str(deno)
        
    return strFract,n,d,nFact,dFact

def exp10 (x):
    return 10**x

# %% Programme

nPrem = -1800
nDern = 1800

nPrem = -900
nDern = 900

baseLog = 10

nomi = 1
deno = 2

b = 1

nbChiffre = 1
nbDecimal = 9

for a in range(nPrem, nDern):
    
    angleDeg = a/10
    angleRad = angleDeg * m.pi/180
    strFract,nomi,deno,x,y = Simplification(a, 1800)
    
    # print(strFract)
    
    L = [angleDeg, MiseEnPage(angleRad,1,9, affichageSigne = True),strFract, None]
    L.append( MiseEnPage( FonctArrondie(angleRad,m.sin,m.asin,12),nbChiffre, nbDecimal, affichageSigne = True))
    L.append( MiseEnPage( FonctArrondie(angleRad,m.cos,m.acos,12),nbChiffre, nbDecimal, affichageSigne = True))
    L.append( MiseEnPage( FonctArrondie(angleRad,m.tan,m.atan,12),3, nbDecimal, affichageSigne = True))
    L.append( None )
    L.append( MiseEnPage( FonctArrondie(abs(m.sin(angleRad)),m.log10,exp10,12),nbChiffre, nbDecimal, affichageSigne = True))
    L.append( MiseEnPage( FonctArrondie(abs(m.cos(angleRad)),m.log10,exp10,12),nbChiffre, nbDecimal, affichageSigne = True))
    L.append( MiseEnPage( FonctArrondie(abs(m.tan(angleRad)),m.log10,exp10,12),3, nbDecimal, affichageSigne = True))

    ws.append(L)
    
    if (a + 1) % 50 == 0:
        ws.append(Intro)
        b += 51
        ws.merge_cells(start_row=b, start_column=2, end_row=b, end_column=3)

wb.save("Section 3 - Trigonometrie/v1 - Trigonometrie/Trigo/sample_3.xlsx")