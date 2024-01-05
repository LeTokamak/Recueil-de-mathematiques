# -*- coding: utf-8 -*-
"""
Created on Wed Jul 20 18:08:49 2022

@author: Tokamak
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Color

wb2 = load_workbook('Table logarithme v3.xlsx')

log = wb2.worksheets[0]

police_gras =        Font(name='Calibri', charset=1, family=2.0, b=True, i=False, strike=None, outline=None, shadow=None, condense=None, color = Color(rgb='FFFFFFFF', indexed=None, auto=None, theme=None, tint=0.0, type='rgb'))
remplisage  = PatternFill(patternType='solid', fgColor = Color(rgb='FF333333', indexed=None, auto=None, theme=None, tint=0.0, type='rgb')) #PatternFill(log.cell(1,3))

for ligne in range (2,9181):
    for col in range(3,13):
        if type(log.cell(ligne, col).value) == int :
            log.cell(ligne, col).font = police_gras
            log.cell(ligne, col).fill = remplisage

          
wb2.save('Table logarithme v5.xlsx')